"""Conversión del Excel del POA institucional al JSON que consume el dashboard.

Estructura esperada del libro (la del archivo POA_INSTITUCIONAL_YYYY.xlsx):

* Una hoja de proyectos (su nombre contiene "PROYECTO").
* Hojas de control/seguimiento que definen el bloque y traen el % de avance
  oficial y el responsable de cada unidad:
    - "CONTROL UNIDADES ADMINISTRATIVAS"  -> Unidad Administrativa
    - "Seguimiento Unidades Académicas"   -> Unidad Académica
    - "Seguimiento Carreras"              -> Carrera
* Después de cada hoja de control vienen las hojas de cada unidad, numeradas
  ("1. UNIDAD DE COMUNICACIÓN", "8.TSEI", ...).

En cada hoja de unidad se busca la fila de encabezados (Estado, Prioridad,
Fecha de entrega, Tarea, Docente Responsable, Descripción, Entregable,
% Ejecutado, Observación...). Las filas cuyo texto empieza con
"OBJETIVO OPERATIVO" son separadores: definen el objetivo de las actividades
que vienen debajo.
"""

from __future__ import annotations

import datetime as dt
import re
import unicodedata

from openpyxl import load_workbook

# Palabras clave por campo de la fila de encabezados.
COLUMNAS = {
    "estado": ("estado",),
    "prioridad": ("prioridad",),
    "fecha": ("fecha", "plazo", "entrega"),
    "tarea": ("tarea", "actividad", "accion"),
    "responsable": ("responsable",),
    "descripcion": ("descripcion", "detalle"),
    "entregable": ("entregable", "producto", "medio de verificacion"),
    "avance": ("% ejecutado", "ejecutado", "avance", "cumplimiento", "porcentaje"),
    "observacion": ("observacion", "observaciones", "nota"),
}

BLOQUES = (
    ("control unidades administrativa", "Unidad Administrativa"),
    ("unidades administrativas", "Unidad Administrativa"),
    ("seguimiento unidades academicas", "Unidad Académica"),
    ("unidades academicas", "Unidad Académica"),
    ("seguimiento carreras", "Carrera"),
    ("carreras", "Carrera"),
)

RE_OBJETIVO = re.compile(r"objetivo\s+(operativo|estrategico|general)?", re.IGNORECASE)
RE_SOLO_ETIQUETA = re.compile(r"^objetivo\s+\w+\s*\d*\s*[:.\-–]?\s*$", re.IGNORECASE)
RE_PREFIJO = re.compile(r"^objetivo\s+\w+\s*\d*\s*[:.\-–]+\s*", re.IGNORECASE)
RE_NUMERO_HOJA = re.compile(r"^\s*(\d+)\s*[.\-)]?\s*(.*)$")


class ExcelPOAError(Exception):
    """Error legible para mostrar al administrador."""


# --------------------------------------------------------------------------- #
# utilidades
# --------------------------------------------------------------------------- #
def _norm(valor) -> str:
    texto = str(valor if valor is not None else "").strip().lower()
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", texto)


def _texto(valor):
    if valor is None:
        return None
    if isinstance(valor, dt.datetime):
        return valor.strftime("%Y-%m-%d")
    if isinstance(valor, dt.date):
        return valor.strftime("%Y-%m-%d")
    texto = re.sub(r"[ \t]+", " ", str(valor)).strip()
    texto = re.sub(r"\s*\n\s*", " ", texto).strip()
    return texto or None


def _titulo_persona(valor):
    texto = _texto(valor)
    if not texto:
        return None
    if texto.isupper():
        return texto.title()
    return texto


def _avance(valor):
    """Devuelve un número entre 0 y 1, o None si la celda está vacía."""
    if valor is None or valor == "":
        return None
    if isinstance(valor, bool):
        return None
    if isinstance(valor, (int, float)):
        numero = float(valor)
    else:
        crudo = str(valor).strip()
        porcentaje = "%" in crudo
        crudo = crudo.replace("%", "").replace(",", ".").strip()
        try:
            numero = float(crudo)
        except ValueError:
            return None
        if porcentaje:
            numero /= 100
    if numero > 1:
        numero /= 100
    return round(max(0.0, min(1.0, numero)), 4)


def _estado(valor):
    if isinstance(valor, bool):
        return valor
    texto = _norm(valor)
    if texto in {"true", "verdadero", "si", "sí", "x", "1", "cumplido"}:
        return True
    if texto in {"false", "falso", "no", "0", ""}:
        return False
    return True


def _es_objetivo(celdas) -> bool:
    for celda in celdas:
        texto = _norm(celda)
        if texto.startswith("objetivo") and RE_OBJETIVO.match(texto):
            return True
    return False


def _texto_objetivo(celdas):
    """Junta la etiqueta del objetivo con su descripción si están separadas."""
    partes = [_texto(c) for c in celdas]
    partes = [p for p in partes if p]
    if not partes:
        return None
    inicio = 0
    for i, parte in enumerate(partes):
        if _norm(parte).startswith("objetivo"):
            inicio = i
            break
    etiqueta = partes[inicio]
    resto = [p for p in partes[inicio + 1 :] if len(p) > 3]
    if RE_SOLO_ETIQUETA.match(etiqueta):
        return resto[0] if resto else etiqueta
    limpio = RE_PREFIJO.sub("", etiqueta).strip()
    limpio = re.sub(r"^\d+[.\-)]\s*", "", limpio).strip()
    if limpio:
        return limpio
    return resto[0] if resto else etiqueta


# --------------------------------------------------------------------------- #
# lectura de hojas
# --------------------------------------------------------------------------- #
def _fila_encabezado(filas, minimo: int = 4):
    for indice, fila in enumerate(filas[:30]):
        normalizadas = [_norm(c) for c in fila]
        aciertos = sum(
            1
            for claves in COLUMNAS.values()
            if any(celda and any(clave in celda for clave in claves) for celda in normalizadas)
        )
        if aciertos >= minimo:
            return indice
    return None


def _mapa_columnas(fila_encabezado):
    mapa: dict[str, int] = {}
    usados: set[int] = set()
    for campo, claves in COLUMNAS.items():
        encontrado = None
        for clave in claves:  # respeta la prioridad de las palabras clave
            for indice, celda in enumerate(fila_encabezado):
                if indice in usados:
                    continue
                texto = _norm(celda)
                if texto and clave in texto:
                    encontrado = indice
                    break
            if encontrado is not None:
                break
        if encontrado is not None:
            mapa[campo] = encontrado
            usados.add(encontrado)
    return mapa


def _titulo_hoja(filas_previas, respaldo):
    candidatos = []
    for fila in filas_previas:
        for celda in fila:
            texto = _texto(celda)
            if texto and len(texto) > 10:
                candidatos.append(texto)
    if not candidatos:
        return respaldo
    return max(candidatos, key=len)


def _hoja_proyectos(hoja):
    filas = [list(f) for f in hoja.iter_rows(values_only=True)]
    indice = _fila_encabezado(filas, minimo=2)
    if indice is None:
        return []
    encabezado = [_norm(c) for c in filas[indice]]

    def columna(*claves, defecto=None):
        for clave in claves:
            for i, celda in enumerate(encabezado):
                if celda and clave in celda:
                    return i
        return defecto

    col_nombre = columna("nombre del proyecto", "proyecto", "nombre", defecto=0)
    col_desc = columna("descripcion", "detalle")
    col_resp = columna("responsable")
    col_avance = columna("% ejecutado", "ejecutado", "avance")
    col_obs = columna("observacion", "resultado")

    def valor(fila, col):
        return _texto(fila[col]) if col is not None and col < len(fila) else None

    proyectos = []
    for fila in filas[indice + 1 :]:
        nombre = valor(fila, col_nombre)
        if not nombre:
            continue
        proyectos.append(
            {
                "nombre": nombre,
                "descripcion": valor(fila, col_desc),
                "responsable": valor(fila, col_resp),
                "avance": _avance(fila[col_avance])
                if col_avance is not None and col_avance < len(fila)
                else None,
                "observacion": valor(fila, col_obs),
            }
        )
    return proyectos


def _hoja_control(hoja):
    """Devuelve {n°: {'nombre','sigla','responsable','avance'}} de una hoja de control."""
    filas = [list(f) for f in hoja.iter_rows(values_only=True)]
    indice = None
    for i, fila in enumerate(filas[:15]):
        normalizadas = [_norm(c) for c in fila]
        if any(c in {"n", "n°", "no", "nro"} for c in normalizadas) and any(
            "unidad" in c or "carrera" in c or "responsable" in c for c in normalizadas
        ):
            indice = i
            break
    if indice is None:
        return {}
    encabezado = [_norm(c) for c in filas[indice]]

    def columna(*claves):
        for clave in claves:
            for i, celda in enumerate(encabezado):
                if celda and clave in celda:
                    return i
        return None

    col_nombre = columna("unidades", "carreras", "unidad", "carrera")
    col_resp = columna("responsab")
    col_sigla = columna("sigla")
    col_avance = columna("% ejecutado", "ejecutado", "avance", "porcentaje")

    # En "Seguimiento Carreras" las columnas de nombre y responsable vienen
    # intercambiadas respecto al encabezado: el nombre real es el que contiene
    # el texto más largo/descriptivo. Se corrige comparando ambas columnas.
    filas_datos = [f for f in filas[indice + 1 :] if any(f)]
    if col_nombre is not None and col_resp is not None:
        def largo(col):
            valores = [len(_texto(f[col]) or "") for f in filas_datos if col < len(f)]
            return sum(valores) / len(valores) if valores else 0

        if largo(col_resp) > largo(col_nombre):
            col_nombre, col_resp = col_resp, col_nombre

    control = {}
    for fila in filas_datos:
        def valor(col):
            return _texto(fila[col]) if col is not None and col < len(fila) else None

        numero_bruto = _texto(fila[0]) if fila else None
        try:
            numero = int(float(str(numero_bruto).replace(",", ".")))
        except (TypeError, ValueError):
            continue
        control[numero] = {
            "nombre": valor(col_nombre),
            "sigla": valor(col_sigla),
            "responsable": _titulo_persona(valor(col_resp)),
            "avance": _avance(fila[col_avance])
            if col_avance is not None and col_avance < len(fila)
            else None,
        }
    return control


def _hoja_unidad(hoja):
    filas = [list(f) for f in hoja.iter_rows(values_only=True)]
    indice = _fila_encabezado(filas)
    if indice is None:
        return None

    mapa = _mapa_columnas(filas[indice])
    if "tarea" not in mapa:
        return None

    actividades = []
    objetivo_actual = None
    for fila in filas[indice + 1 :]:
        if not any(c not in (None, "") for c in fila):
            continue

        def valor(campo):
            col = mapa.get(campo)
            if col is None or col >= len(fila):
                return None
            return fila[col]

        if _es_objetivo(fila):
            objetivo_actual = _texto_objetivo(fila)
            continue

        tarea = _texto(valor("tarea"))
        if not tarea:
            continue
        if _norm(tarea) in {"tarea", "actividad"}:
            continue

        actividades.append(
            {
                "objetivo": objetivo_actual,
                "tarea": tarea,
                "prioridad": _texto(valor("prioridad")),
                "fecha": _texto(valor("fecha")),
                "responsable": _texto(valor("responsable")),
                "entregable": _texto(valor("entregable")),
                "descripcion": _texto(valor("descripcion")),
                "avance": _avance(valor("avance")),
                "observacion": _texto(valor("observacion")),
                "estado": _estado(valor("estado")),
            }
        )

    if not actividades:
        return None

    con_avance = [a["avance"] for a in actividades if a["avance"] is not None]
    promedio = round(sum(con_avance) / len(con_avance), 4) if con_avance else 0.0
    nombre = re.sub(r"^\s*\d+\s*[.\-)]?\s*", "", hoja.title).strip() or hoja.title

    return {
        "id": hoja.title,
        "nombre": nombre,
        "titulo": _titulo_hoja(filas[:indice], nombre),
        "categoria": "Unidad Administrativa",
        "responsable": None,
        "avance": promedio,
        "avanceActividades": promedio,
        "totalActividades": len(actividades),
        "actividades": actividades,
    }


# --------------------------------------------------------------------------- #
# conversión completa
# --------------------------------------------------------------------------- #
def _categoria_de_control(titulo_hoja):
    texto = _norm(titulo_hoja)
    for clave, categoria in BLOQUES:
        if clave in texto:
            return categoria
    return None


def _es_hoja_ignorada(titulo_hoja):
    texto = _norm(titulo_hoja)
    if texto in {"poa", "poa 2026", "resumen", "instructivo"}:
        return True
    return texto.startswith("estrategia") or texto.startswith("poa institucional")


def poa_desde_excel(archivo, anio: int) -> dict:
    """Convierte un libro de Excel del POA al diccionario del dashboard."""
    try:
        libro = load_workbook(archivo, data_only=True, read_only=True)
    except Exception as exc:  # pragma: no cover - depende del archivo subido
        raise ExcelPOAError(f"No se pudo leer el archivo de Excel: {exc}") from exc

    unidades: list[dict] = []
    proyectos: list[dict] = []
    omitidas: list[str] = []

    categoria_actual = "Unidad Administrativa"
    control_actual: dict[int, dict] = {}

    for hoja in libro.worksheets:
        titulo = hoja.title

        if "proyecto" in _norm(titulo) and "operativo" not in _norm(titulo):
            proyectos.extend(_hoja_proyectos(hoja))
            continue

        categoria = _categoria_de_control(titulo)
        if categoria:
            categoria_actual = categoria
            control_actual = _hoja_control(hoja)
            continue

        if _es_hoja_ignorada(titulo):
            omitidas.append(titulo)
            continue

        unidad = _hoja_unidad(hoja)
        if not unidad:
            omitidas.append(titulo)
            continue

        unidad["categoria"] = categoria_actual

        # Enlaza con la fila de control: primero por sigla, luego por número.
        coincidencia = None
        clave_hoja = _norm(unidad["nombre"])
        for datos in control_actual.values():
            sigla = _norm(datos.get("sigla"))
            if sigla and sigla == clave_hoja:
                coincidencia = datos
                break
        if coincidencia is None:
            m = RE_NUMERO_HOJA.match(titulo)
            if m:
                coincidencia = control_actual.get(int(m.group(1)))

        if coincidencia:
            unidad["responsable"] = coincidencia.get("responsable")
            if coincidencia.get("nombre"):
                unidad["nombreOficial"] = coincidencia["nombre"]
            if coincidencia.get("avance") is not None:
                unidad["avance"] = coincidencia["avance"]

        unidades.append(unidad)

    libro.close()

    if not unidades:
        raise ExcelPOAError(
            "No se encontró ninguna hoja con actividades. Revisa que cada hoja tenga una fila "
            "de encabezados con columnas como Estado, Prioridad, Fecha de entrega, Tarea, "
            "Entregable y % Ejecutado."
        )

    return {
        "anio": anio,
        "unidades": unidades,
        "proyectos": proyectos,
        "_omitidas": omitidas,
    }
